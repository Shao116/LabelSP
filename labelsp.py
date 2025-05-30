import os
import sys
import csv
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QPushButton, QFileDialog, QLabel, QMessageBox, QGroupBox,
                             QFrame, QProgressDialog, QSplashScreen)
from PyQt5.QtGui import (QPixmap, QImage, QPainter, QColor, QPen, QCursor, QFont)
from PyQt5.QtCore import (Qt, QPoint, QPointF, QRectF, QTimer, QSettings)
from PyQt5.QtWidgets import QGraphicsView, QGraphicsScene, QGraphicsPixmapItem, QGraphicsItem

# Shaozetong produced
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    import rasterio
    from rasterio.transform import Affine
    from rasterio.crs import CRS
    import pyproj
except ImportError:
    print("Installing required modules...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "rasterio", "pyproj"])
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    import rasterio
    from rasterio.transform import Affine
    from rasterio.crs import CRS
    import pyproj

if hasattr(Qt, 'AA_EnableHighDpiScaling'):
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

app = QApplication(sys.argv)

# Show splash screen
splash_pix = QPixmap('1.jpg') if os.path.exists('1.jpg') else QPixmap(400, 300)
splash_pix = splash_pix.scaled(800, 600, Qt.KeepAspectRatio, Qt.SmoothTransformation)
splash = QSplashScreen(splash_pix, Qt.WindowStaysOnTopHint)
splash.show()
QTimer.singleShot(3000, splash.close)  # Display for 3 seconds

class CrosshairItem(QGraphicsItem):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setZValue(1000)
        self.cross_pos = QPoint(0, 0)
        self.pen = QPen(QColor(0, 255, 0), 1)
        self.pen.setCosmetic(True)
        self.setAcceptHoverEvents(True)
        
    def boundingRect(self):
        return QRectF(-20, -20, 40, 40)
        
    def paint(self, painter, option, widget):
        painter.setPen(self.pen)
        painter.drawLine(QPointF(-20, 0), QPointF(20, 0))
        painter.drawLine(QPointF(0, -20), QPointF(0, 20))
        painter.drawPoint(QPointF(0, 0))
        
    def updatePosition(self, pos):
        self.setPos(pos)
        self.cross_pos = pos

class AnnotationPoint(QGraphicsItem):
    def __init__(self, pos, parent=None):
        super().__init__(parent)
        self.setPos(pos)
        self.setZValue(100)
        self.pen = QPen(QColor(255, 0, 0), 1)
        self.pen.setCosmetic(True)
        self.is_selected = False
        
    def boundingRect(self):
        return QRectF(-5, -5, 10, 10)
        
    def paint(self, painter, option, widget):
        if self.is_selected:
            painter.setPen(QPen(QColor(0, 0, 255), 1))
        else:
            painter.setPen(self.pen)
        painter.drawLine(QPointF(-5, 0), QPointF(5, 0))
        painter.drawLine(QPointF(0, -5), QPointF(0, 5))
        
    def setSelected(self, selected):
        self.is_selected = selected
        self.update()

class ImageViewer(QGraphicsView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.scene = QGraphicsScene(self)
        self.setScene(self.scene)

        self.pixmap_item = QGraphicsPixmapItem()
        self.scene.addItem(self.pixmap_item)

        self.setRenderHint(QPainter.Antialiasing)
        self.setRenderHint(QPainter.SmoothPixmapTransform)
        
        self.mode = "click"
        self.setCursor(Qt.CrossCursor)

        self.crosshair = CrosshairItem()
        self.scene.addItem(self.crosshair)
        self.crosshair.hide()

        self.scale_factor = 1.0
        self.annotations = []
        self.temp_annotations = []
        
        self.select_start = None
        self.select_rect = None
        self.selecting = False

        self.setMouseTracking(True)
        self.setAcceptDrops(True)
        
        # TIFF related attributes
        # Shaozetong produced
        self.tif_file = None
        self.transform = None
        self.crs = None
        self.transformer = None

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                self.parent.load_image_with_progress(file_path)
                break

    def load_image(self, image_path):
        if not os.path.exists(image_path):
            QMessageBox.warning(self, "警告", "图像文件未找到!")
            return False
        
        try:
            progress = QProgressDialog("正在加载图片...", None, 0, 0, self)
            progress.setWindowTitle("请稍候")
            progress.setWindowModality(Qt.WindowModal)
            progress.setCancelButton(None)
            progress.show()
            
            QApplication.processEvents()
            
            # Reset TIFF attributes
            self.tif_file = None
            self.transform = None
            self.crs = None
            self.transformer = None
            
            # For TIFF files, try to read geospatial info
            if image_path.lower().endswith(('.tif', '.tiff')):
                try:
                    self.tif_file = rasterio.open(image_path)
                    self.transform = self.tif_file.transform
                    self.crs = self.tif_file.crs
                    
                    # Create coordinate transformer if CRS is not WGS84
                    if self.crs and not self.crs.is_geographic:
                        try:
                            wgs84 = CRS.from_epsg(4326)  # WGS84
                            self.transformer = pyproj.Transformer.from_crs(self.crs, wgs84, always_xy=True)
                        except Exception as e:
                            print(f"Warning: Could not create coordinate transformer: {str(e)}")
                except Exception as e:
                    print(f"Warning: Could not read geospatial info from TIFF: {str(e)}")
            
            image = QImage(image_path)
            if image.isNull():
                progress.close()
                QMessageBox.warning(self, "警告", "加载图像失败!")
                return False
            
            self.pixmap_item.setPixmap(QPixmap.fromImage(image))
            self.scene.setSceneRect(self.pixmap_item.boundingRect())
            self.fitInView(self.pixmap_item, Qt.KeepAspectRatio)
            self.scale_factor = 1.0
            self.annotations = []
            self.temp_annotations = []
            self.crosshair.show()
            
            progress.close()
            return True
        except Exception as e:
            if 'progress' in locals():
                progress.close()
            QMessageBox.warning(self, "错误", f"加载图像失败: {str(e)}")
            return False

# Shaozetong produced
    def wheelEvent(self, event):
        zoom_factor = 1.2
        old_pos = self.mapToScene(event.pos())

        if event.angleDelta().y() > 0:
            self.scale(zoom_factor, zoom_factor)
            self.scale_factor *= zoom_factor
        else:
            self.scale(1 / zoom_factor, 1 / zoom_factor)
            self.scale_factor /= zoom_factor

        new_pos = self.mapToScene(event.pos())
        delta = new_pos - old_pos
        self.translate(delta.x(), delta.y())

    def mouseMoveEvent(self, event):
        scene_pos = self.mapToScene(event.pos())
        self.crosshair.updatePosition(scene_pos)
        
        if self.pixmap_item.pixmap() and not self.pixmap_item.pixmap().isNull():
            img_width = self.pixmap_item.pixmap().width()
            img_height = self.pixmap_item.pixmap().height()
            
            x = scene_pos.x()
            y = scene_pos.y()
            normalized_x = x / img_width
            normalized_y = y / img_height
            
            coord_text = f"X: {x:.1f}, Y: {y:.1f} (Norm: {normalized_x:.3f}, {normalized_y:.3f})"
            
            # Add lon/lat if available
            if self.transform is not None:
                try:
                    lon, lat = self.pixel_to_coords(x, y)
                    if self.transformer is not None:
                        # Convert to WGS84 if not already
                        lon, lat = self.transformer.transform(lon, lat)
                    coord_text += f" | Lon: {lon:.6f}, Lat: {lat:.6f}"
                except Exception as e:
                    print(f"Error converting coordinates: {str(e)}")
            
            self.parent.update_coord_label(coord_text)
        
        if self.selecting and self.mode == "select":
            end_pos = self.mapToScene(event.pos())
            if self.select_rect:
                self.scene.removeItem(self.select_rect)
            
            rect = QRectF(self.select_start, end_pos).normalized()
            self.select_rect = self.scene.addRect(rect, QPen(Qt.blue, 1, Qt.DashLine))
            self.select_rect.setZValue(1000)
            
            for point in self.annotations + self.temp_annotations:
                if rect.contains(point.pos()):
                    point.setSelected(True)
                else:
                    point.setSelected(False)
        
        super().mouseMoveEvent(event)

    def pixel_to_coords(self, x, y):
        """Convert pixel coordinates to geographic coordinates"""
        if self.transform is None:
            raise ValueError("No geospatial transform available")
        
        # Rasterio uses (row, col) order, so we need to flip x and y
        lon, lat = self.transform * (x, y)
        return lon, lat

    def mousePressEvent(self, event):
        if self.mode == "click":
            if event.button() == Qt.RightButton or (event.button() == Qt.LeftButton and event.modifiers() == Qt.KeyboardModifier.NoModifier):
                scene_pos = self.mapToScene(event.pos())
                if self.pixmap_item.contains(scene_pos):
                    point = AnnotationPoint(scene_pos)
                    self.scene.addItem(point)
                    self.temp_annotations.append(point)
                    
                    # Get coordinates for status message
                    x, y = scene_pos.x(), scene_pos.y()
                    status_msg = f"添加标注点位置: {round(x, 1)}, {round(y, 1)}"
                    
                    if self.transform is not None:
                        try:
                            lon, lat = self.pixel_to_coords(x, y)
                            if self.transformer is not None:
                                lon, lat = self.transformer.transform(lon, lat)
                            status_msg += f" (Lon: {lon:.6f}, Lat: {lat:.6f})"
                        except Exception as e:
                            print(f"Error converting coordinates: {str(e)}")
                    
                    self.parent.update_status_bar(status_msg)
        elif self.mode == "select" and event.button() == Qt.LeftButton:
            self.select_start = self.mapToScene(event.pos())
            self.selecting = True
            for point in self.annotations + self.temp_annotations:
                point.setSelected(False)
        
        super().mousePressEvent(event)

# Shaozetong produced
    def mouseReleaseEvent(self, event):
        if self.selecting and self.mode == "select":
            self.selecting = False
            if self.select_rect:
                end_pos = self.mapToScene(event.pos())
                select_rect = QRectF(self.select_start, end_pos).normalized()
                
                selected_points = []
                for point in self.annotations + self.temp_annotations:
                    if select_rect.contains(point.pos()):
                        selected_points.append(point)
                
                if selected_points:
                    reply = QMessageBox.question(
                        self, "删除标注",
                        f"删除选中的{len(selected_points)}个标注?",
                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                    
                    if reply == QMessageBox.Yes:
                        for point in selected_points:
                            if point in self.annotations:
                                self.annotations.remove(point)
                            elif point in self.temp_annotations:
                                self.temp_annotations.remove(point)
                            self.scene.removeItem(point)
                        self.parent.update_status_bar(f"已删除{len(selected_points)}个标注")
                
                if self.select_rect:
                    self.scene.removeItem(self.select_rect)
                    self.select_rect = None
                
                for point in self.annotations + self.temp_annotations:
                    point.setSelected(False)
        
        super().mouseReleaseEvent(event)

    def keyPressEvent(self, event):
        if self.mode == "click":
            if event.key() == Qt.Key_Space:
                scene_pos = self.mapToScene(self.viewport().mapFromGlobal(QCursor.pos()))
                if self.pixmap_item.contains(scene_pos):
                    point = AnnotationPoint(scene_pos)
                    self.scene.addItem(point)
                    self.temp_annotations.append(point)
                    
                    # Get coordinates for status message
                    x, y = scene_pos.x(), scene_pos.y()
                    status_msg = f"添加标注点位置: {round(x, 1)}, {round(y, 1)}"
                    
                    if self.transform is not None:
                        try:
                            lon, lat = self.pixel_to_coords(x, y)
                            if self.transformer is not None:
                                lon, lat = self.transformer.transform(lon, lat)
                            status_msg += f" (Lon: {lon:.6f}, Lat: {lat:.6f})"
                        except Exception as e:
                            print(f"Error converting coordinates: {str(e)}")
                    
                    self.parent.update_status_bar(status_msg)
        elif event.key() == Qt.Key_Escape and self.mode == "select":
            if self.select_rect:
                self.scene.removeItem(self.select_rect)
                self.select_rect = None
            for point in self.annotations + self.temp_annotations:
                point.setSelected(False)
            self.selecting = False
        
        super().keyPressEvent(event)

    def set_click_mode(self):
        self.mode = "click"
        self.setDragMode(QGraphicsView.NoDrag)
        self.setCursor(Qt.CrossCursor)
        self.parent.update_status_bar("已切换到点击模式")

    def set_drag_mode(self):
        self.mode = "drag"
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setCursor(Qt.OpenHandCursor)
        self.parent.update_status_bar("已切换到拖动模式")
        
    def set_select_mode(self):
        self.mode = "select"
        self.setDragMode(QGraphicsView.NoDrag)
        self.setCursor(Qt.CrossCursor)
        self.parent.update_status_bar("已切换到选择模式")

    def zoom_in(self):
        self.scale(1.2, 1.2)
        self.scale_factor *= 1.2

    def zoom_out(self):
        self.scale(1 / 1.2, 1 / 1.2)
        self.scale_factor /= 1.2

    def reset_zoom(self):
        self.resetTransform()
        self.fitInView(self.pixmap_item, Qt.KeepAspectRatio)
        self.scale_factor = 1.0
# Shaozetong produced
    def undo_annotation(self):
        if self.temp_annotations:
            point = self.temp_annotations.pop()
            self.scene.removeItem(point)
            pos = point.pos()
            self.parent.update_status_bar(f"移除标注点位置: {round(pos.x(), 1)}, {round(pos.y(), 1)}")
        elif self.annotations:
            point = self.annotations.pop()
            self.scene.removeItem(point)
            pos = point.pos()
            self.parent.update_status_bar(f"移除标注点位置: {round(pos.x(), 1)}, {round(pos.y(), 1)}")
        else:
            self.parent.update_status_bar("没有可撤销的标注")

    def confirm_annotations(self):
        self.annotations.extend(self.temp_annotations)
        self.temp_annotations = []
        self.parent.update_status_bar("标注已确认")

    def clear_annotations(self):
        for point in self.annotations + self.temp_annotations:
            self.scene.removeItem(point)
        self.annotations = []
        self.temp_annotations = []
        self.parent.update_status_bar("所有标注已清除")

    def get_annotations(self):
        return self.annotations + self.temp_annotations

    def get_normalized_annotations(self):
        if not self.pixmap_item.pixmap().isNull():
            img_width = self.pixmap_item.pixmap().width()
            img_height = self.pixmap_item.pixmap().height()
            
            normalized = []
            for point in self.get_annotations():
                pos = point.pos()
                x = pos.x() / img_width
                y = pos.y() / img_height
                normalized.append((x, y))
            
            return normalized
        return []

class ImageAnnotationTool(QMainWindow):
    def __init__(self):
        super().__init__()
        
        self.settings = QSettings("ImageAnnotationTool", "ImageAnnotationTool")
        self.init_ui()
        self.setWindowTitle("终极标注V3.0 LTS (支持地理坐标)")
        self.resize(1200, 800)
        
# Shaozetong produced
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F0F0F0;
            }
            QGroupBox {
                background-color: #FFFFFF;
                border: 1px solid #D4D4D4;
                border-radius: 3px;
                margin-top: 10px;
                padding-top: 15px;
                font-size: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 7px;
                padding: 0 3px;
            }
            QPushButton {
                background-color: #F0F0F0;
                border: 1px solid #D4D4D4;
                border-radius: 3px;
                padding: 5px 10px;
                min-width: 80px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #E5E5E5;
            }
            QPushButton:pressed {
                background-color: #D4D4D4;
            }
            QLabel {
                font-size: 12px;
            }
        """)
        

        last_file = self.settings.value("last_image", "")
        if last_file and os.path.exists(last_file):
            reply = QMessageBox.question(
                self, "打开上次文件",
                "是否打开上次的文件?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            
            if reply == QMessageBox.Yes:
                self.load_image_with_progress(last_file)

    def init_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        self.main_layout = QHBoxLayout()
        self.main_layout.setContentsMargins(5, 5, 5, 5)
        self.central_widget.setLayout(self.main_layout)
        

        self.left_toolbar = QGroupBox("工具")
        self.left_toolbar.setFixedWidth(150)
        self.left_toolbar_layout = QVBoxLayout()
        self.left_toolbar_layout.setContentsMargins(10, 15, 10, 10)
        self.left_toolbar_layout.setSpacing(8)
        self.left_toolbar.setLayout(self.left_toolbar_layout)
        self.main_layout.addWidget(self.left_toolbar)
        
# Shaozetong produced
        self.right_panel = QWidget()
        self.right_layout = QVBoxLayout()
        self.right_layout.setContentsMargins(5, 5, 5, 5)
        self.right_panel.setLayout(self.right_layout)
        self.main_layout.addWidget(self.right_panel)
        

        self.image_viewer = ImageViewer(self)
        self.right_layout.addWidget(self.image_viewer)
        

        self.status_bar = QHBoxLayout()
        self.status_bar.setContentsMargins(5, 5, 5, 5)
        self.coord_label = QLabel("X: 0, Y: 0 (Norm: 0, 0)")
        self.coord_label.setFont(QFont("Arial", 10))
        self.status_label = QLabel("就绪")
        self.status_label.setFont(QFont("Arial", 10))
        self.status_bar.addWidget(self.coord_label)
        self.status_bar.addStretch()
        self.status_bar.addWidget(self.status_label)
        self.right_layout.addLayout(self.status_bar)
        
        self.create_toolbar()
        self.setAcceptDrops(True)

    def create_toolbar(self):

        button_style = "QPushButton { text-align: left; padding: 5px 10px; font-size: 12px; }"
        
        self.open_button = QPushButton("打开图像")
        self.open_button.setStyleSheet(button_style)
        self.open_button.clicked.connect(self.open_image)
        self.left_toolbar_layout.addWidget(self.open_button)

        self.import_button = QPushButton("导入标注")
        self.import_button.setStyleSheet(button_style)
        self.import_button.clicked.connect(self.import_annotations)
        self.left_toolbar_layout.addWidget(self.import_button)

        self.export_button = QPushButton("导出标注")
        self.export_button.setStyleSheet(button_style)
        self.export_button.clicked.connect(self.export_annotations)
        self.left_toolbar_layout.addWidget(self.export_button)


        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        self.left_toolbar_layout.addWidget(separator)

        self.click_mode_button = QPushButton("点击模式")
        self.click_mode_button.setStyleSheet(button_style)
        self.click_mode_button.clicked.connect(self.image_viewer.set_click_mode)
        self.left_toolbar_layout.addWidget(self.click_mode_button)

        self.drag_mode_button = QPushButton("拖动模式")
        self.drag_mode_button.setStyleSheet(button_style)
        self.drag_mode_button.clicked.connect(self.image_viewer.set_drag_mode)
        self.left_toolbar_layout.addWidget(self.drag_mode_button)

        self.select_mode_button = QPushButton("选择模式")
        self.select_mode_button.setStyleSheet(button_style)
        self.select_mode_button.clicked.connect(self.image_viewer.set_select_mode)
        self.left_toolbar_layout.addWidget(self.select_mode_button)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        self.left_toolbar_layout.addWidget(separator)

        self.zoom_in_button = QPushButton("放大 (+)")
        self.zoom_in_button.setStyleSheet(button_style)
        self.zoom_in_button.clicked.connect(self.image_viewer.zoom_in)
        self.left_toolbar_layout.addWidget(self.zoom_in_button)

        self.zoom_out_button = QPushButton("缩小 (-)")
        self.zoom_out_button.setStyleSheet(button_style)
        self.zoom_out_button.clicked.connect(self.image_viewer.zoom_out)
        self.left_toolbar_layout.addWidget(self.zoom_out_button)

        self.reset_zoom_button = QPushButton("重置缩放")
        self.reset_zoom_button.setStyleSheet(button_style)
        self.reset_zoom_button.clicked.connect(self.image_viewer.reset_zoom)
        self.left_toolbar_layout.addWidget(self.reset_zoom_button)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        self.left_toolbar_layout.addWidget(separator)

        self.confirm_button = QPushButton("确认标注")
        self.confirm_button.setStyleSheet(button_style)
        self.confirm_button.clicked.connect(self.image_viewer.confirm_annotations)
        self.left_toolbar_layout.addWidget(self.confirm_button)

        self.clear_button = QPushButton("清除标注")
        self.clear_button.setStyleSheet(button_style)
        self.clear_button.clicked.connect(self.image_viewer.clear_annotations)
        self.left_toolbar_layout.addWidget(self.clear_button)
# Shaozetong produced
        self.left_toolbar_layout.addStretch()

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                self.load_image_with_progress(file_path)
                break

    def load_image_with_progress(self, file_path):
        progress = QProgressDialog("正在加载图片...", None, 0, 0, self)
        progress.setWindowTitle("请稍候")
        progress.setWindowModality(Qt.WindowModal)
        progress.setCancelButton(None)
        progress.show()
        
        QApplication.processEvents()
        
        QTimer.singleShot(100, lambda: self._load_image_after_delay(file_path, progress))

    def _load_image_after_delay(self, file_path, progress):
        try:
            if self.image_viewer.load_image(file_path):
                self.update_status_bar(f"已加载: {os.path.basename(file_path)}")
                self.settings.setValue("last_image", file_path)
                
# Shaozetong produced
                if file_path.lower().endswith(('.tif', '.tiff')) and self.image_viewer.transform is not None:
                    crs_info = str(self.image_viewer.crs) if self.image_viewer.crs else "未知"
                    self.update_status_bar(f"已加载: {os.path.basename(file_path)} (CRS: {crs_info})")
        finally:
            progress.close()

    def open_image(self):
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("打开图像")
        file_dialog.setNameFilter("Images (*.png *.jpg *.jpeg *.bmp *.tif *.tiff)")
        
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            self.load_image_with_progress(file_path)

    def import_annotations(self):
        if not self.image_viewer.pixmap_item.pixmap() or self.image_viewer.pixmap_item.pixmap().isNull():
            QMessageBox.warning(self, "警告", "请先加载图像!")
            return
        
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("导入标注")
        file_dialog.setNameFilter("Excel Files (*.xlsx);;CSV Files (*.csv)")
        
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            try:
                if file_path.lower().endswith('.xlsx'):
                    self.import_from_xlsx(file_path)
                elif file_path.lower().endswith('.csv'):
                    self.import_from_csv(file_path)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"导入失败: {str(e)}")

    def import_from_xlsx(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
# Shaozetong produced       
        self.image_viewer.clear_annotations()
        
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            if len(row) >= 2:
                x, y = row[0], row[1]
                if isinstance(x, (int, float)) and isinstance(y, (int, float)):
                    point = AnnotationPoint(QPointF(x, y))
                    self.image_viewer.scene.addItem(point)
                    self.image_viewer.annotations.append(point)
        
        self.update_status_bar(f"从Excel导入 {len(self.image_viewer.annotations)} 个标注点")

    def import_from_csv(self, file_path):
        self.image_viewer.clear_annotations()
        
        with open(file_path, 'r') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip header
            for row in reader:
                if len(row) >= 2:
                    try:
                        x, y = float(row[0]), float(row[1])
                        point = AnnotationPoint(QPointF(x, y))
                        self.image_viewer.scene.addItem(point)
                        self.image_viewer.annotations.append(point)
                    except ValueError:
                        continue
        
        self.update_status_bar(f"从CSV导入 {len(self.image_viewer.annotations)} 个标注点")

    def export_annotations(self):
        if not self.image_viewer.get_annotations():
            QMessageBox.warning(self, "警告", "没有标注可导出!")
            return
        
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("导出标注")
        file_dialog.setAcceptMode(QFileDialog.AcceptSave)
        file_dialog.setNameFilter("Excel Files (*.xlsx)")
        file_dialog.setDefaultSuffix("xlsx")
        
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            try:
                self.export_to_xlsx(file_path)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")
# Shaozetong produced
    def export_to_xlsx(self, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Annotations"
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        
        # Write headers
        headers = ["X", "Y", "Normalized X", "Normalized Y"]
        if self.image_viewer.transform is not None:
            headers.extend(["Longitude", "Latitude"])
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        img_width = self.image_viewer.pixmap_item.pixmap().width()
        img_height = self.image_viewer.pixmap_item.pixmap().height()
        
        for i, point in enumerate(self.image_viewer.get_annotations(), start=2):
            pos = point.pos()
            x, y = pos.x(), pos.y()
            norm_x = x / img_width
            norm_y = y / img_height
            
            ws.cell(row=i, column=1, value=x)
            ws.cell(row=i, column=2, value=y)
            ws.cell(row=i, column=3, value=norm_x)
            ws.cell(row=i, column=4, value=norm_y)
            
            # Add lon/lat if available
            if self.image_viewer.transform is not None:
                try:
                    lon, lat = self.image_viewer.pixel_to_coords(x, y)
                    if self.image_viewer.transformer is not None:
                        # Convert to WGS84 if not already
                        lon, lat = self.image_viewer.transformer.transform(lon, lat)
                    ws.cell(row=i, column=5, value=lon)
                    ws.cell(row=i, column=6, value=lat)
                except Exception as e:
                    print(f"Error converting coordinates: {str(e)}")
            
            # Apply border to data cells
            for col in range(1, len(headers)+1):
                ws.cell(row=i, column=col).border = thin_border
        
        # Auto-size columns
        for col in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        wb.save(file_path)
        self.update_status_bar(f"标注已导出为Excel: {os.path.basename(file_path)}")

    def update_status_bar(self, message):
        self.status_label.setText(message)

    def update_coord_label(self, coord_text):
        self.coord_label.setText(coord_text)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Z and event.modifiers() == Qt.ControlModifier:
            self.image_viewer.undo_annotation()
        else:
            super().keyPressEvent(event)

    def closeEvent(self, event):
        if self.image_viewer.get_annotations():
            reply = QMessageBox.question(
                self, "退出",
                "您有未保存的标注。确定要退出吗?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            
            if reply == QMessageBox.No:
                event.ignore()
                return
        
        if self.image_viewer.tif_file is not None:
            self.image_viewer.tif_file.close()
        
        event.accept()

if __name__ == "__main__":
    window = ImageAnnotationTool()
    window.show()
    sys.exit(app.exec_())# Shaozetong produced